from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


REQUIRED_SHEETS = {
    "README",
    "Workflow Overview",
    "Workflow Steps",
    "Policy Controls",
    "Data Dictionary",
    "Sample Records",
}

OPTIONAL_SHEETS = {
    "Goals & Metrics",
    "Target Systems",
    "Allowed Values",
}

REQUIRED_COLUMNS = {
    "Workflow Overview": [
        "section",
        "response",
        "required",
        "guidance",
    ],
    "Workflow Steps": [
        "step_id",
        "step_name",
        "sequence",
        "owner_role",
        "trigger_or_input",
        "activity",
        "decision_or_rule",
        "systems_used",
        "data_used",
        "output",
        "exceptions_or_escalations",
        "current_pain_points",
    ],
    "Policy Controls": [
        "control_id",
        "control_name",
        "control_type",
        "applies_to_steps",
        "requirement",
        "approval_required",
        "approval_role",
        "evidence_required",
        "write_action_allowed",
        "retention_requirement",
        "source_reference",
    ],
    "Data Dictionary": [
        "field_name",
        "business_meaning",
        "source_system",
        "data_category",
        "required_for_workflow",
        "model_context_allowed",
        "redaction_required",
        "allowed_values",
        "used_in_steps",
        "notes",
    ],
    "Sample Records": [
        "record_id",
    ],
    "Goals & Metrics": [
        "section",
        "response",
        "required",
        "guidance",
    ],
    "Target Systems": [
        "system_name",
        "system_type",
        "read_access_possible",
        "write_access_possible",
        "owner_role",
        "authentication_method",
        "notes",
    ],
}

REQUIRED_OVERVIEW_SECTIONS = {
    "Workflow Name",
    "Business Purpose",
    "Workflow Trigger",
    "Workflow Completion Criteria",
    "Primary Participants",
    "Systems Involved",
    "Current Pain Points",
    "AI Goals",
    "AI No-Go Areas",
    "Known Constraints",
}

DATA_CATEGORIES = {
    "Public",
    "Internal",
    "Confidential",
    "PII",
    "Financial",
    "Security",
    "Sensitive Access",
    "Credential/Secret",
    "Regulated",
    "Derived Metadata",
    "Unknown",
}

CONTROL_TYPES = {
    "approval",
    "audit",
    "data",
    "security",
    "policy",
    "retention",
    "operational",
    "exception",
}

SYSTEM_TYPES = {
    "workflow_system",
    "source_system",
    "execution_system",
    "reporting_system",
    "identity_system",
    "data_store",
    "other",
}

BOOLEAN_FIELDS = {
    "required",
    "approval_required",
    "write_action_allowed",
    "required_for_workflow",
    "model_context_allowed",
    "redaction_required",
    "read_access_possible",
    "write_access_possible",
}

SECRET_OR_SENSITIVE_PATTERN = re.compile(
    r"(password|passwd|secret|token|api[_ -]?key|private[_ -]?key|bearer\s+[a-z0-9._\-]+|-----BEGIN)",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class PacketValidationIssue:
    severity: str
    code: str
    location: str
    message: str


@dataclass(frozen=True)
class PacketValidationResult:
    path: str
    valid: bool
    error_count: int
    warning_count: int
    issues: list[PacketValidationIssue]


def validate_workflow_packet_v1(
    path: str | Path,
    *,
    validate_required_content: bool = True,
) -> PacketValidationResult:
    packet_path = Path(path)
    issues: list[PacketValidationIssue] = []

    if not packet_path.exists():
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="packet.file_not_found",
                location=str(packet_path),
                message="Workflow packet workbook was not found.",
            )
        )
        return _result(packet_path, issues)

    if packet_path.suffix.lower() != ".xlsx":
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="packet.invalid_extension",
                location=str(packet_path),
                message="Workflow Packet v1 must be an .xlsx workbook.",
            )
        )
        return _result(packet_path, issues)

    try:
        workbook = load_workbook(packet_path, read_only=True, data_only=True)
    except Exception as exc:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="packet.cannot_open",
                location=str(packet_path),
                message=f"Workbook could not be opened: {exc}",
            )
        )
        return _result(packet_path, issues)

    _validate_required_sheets(workbook, issues)

    if _has_blocking_sheet_errors(issues):
        return _result(packet_path, issues)

    sheet_rows = {
        sheet_name: _read_sheet_rows(workbook, sheet_name, issues)
        for sheet_name in workbook.sheetnames
        if sheet_name in REQUIRED_COLUMNS
    }

    _validate_required_columns(workbook, issues)

    _validate_workflow_overview(
        rows=sheet_rows.get("Workflow Overview", []),
        issues=issues,
        validate_required_content=validate_required_content,
    )

    step_ids = _validate_workflow_steps(sheet_rows.get("Workflow Steps", []), issues)
    _validate_policy_controls(
        rows=sheet_rows.get("Policy Controls", []),
        valid_step_ids=step_ids,
        issues=issues,
    )

    data_fields = _validate_data_dictionary(
        rows=sheet_rows.get("Data Dictionary", []),
        valid_step_ids=step_ids,
        issues=issues,
    )

    _validate_sample_records(
        rows=sheet_rows.get("Sample Records", []),
        valid_data_fields=data_fields,
        issues=issues,
    )

    if "Target Systems" in workbook.sheetnames:
        _validate_target_systems(
            rows=sheet_rows.get("Target Systems", []),
            issues=issues,
        )
    else:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="packet.optional_target_systems_missing",
                location="Target Systems",
                message="Target Systems sheet is missing. Integration recommendations may be less specific.",
            )
        )

    if "Goals & Metrics" not in workbook.sheetnames:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="packet.optional_goals_metrics_missing",
                location="Goals & Metrics",
                message="Goals & Metrics sheet is missing. Value hypotheses will be directional only.",
            )
        )

    return _result(packet_path, issues)


def _validate_required_sheets(
    workbook: Workbook,
    issues: list[PacketValidationIssue],
) -> None:
    available = set(workbook.sheetnames)

    for sheet_name in sorted(REQUIRED_SHEETS):
        if sheet_name not in available:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="packet.required_sheet_missing",
                    location=sheet_name,
                    message=f"Required sheet '{sheet_name}' is missing.",
                )
            )


def _validate_required_columns(
    workbook: Workbook,
    issues: list[PacketValidationIssue],
) -> None:
    for sheet_name, required_columns in REQUIRED_COLUMNS.items():
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        headers = _read_headers(sheet)

        for column_name in required_columns:
            if column_name not in headers:
                issues.append(
                    PacketValidationIssue(
                        severity="error",
                        code="sheet.required_column_missing",
                        location=f"{sheet_name}.{column_name}",
                        message=f"Required column '{column_name}' is missing from sheet '{sheet_name}'.",
                    )
                )


def _validate_workflow_overview(
    *,
    rows: list[dict[str, Any]],
    issues: list[PacketValidationIssue],
    validate_required_content: bool,
) -> None:
    sections = {_text(row.get("section")): row for row in rows}

    for section in sorted(REQUIRED_OVERVIEW_SECTIONS):
        row = sections.get(section)

        if row is None:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="overview.required_section_missing",
                    location=f"Workflow Overview.{section}",
                    message=f"Required workflow overview section '{section}' is missing.",
                )
            )
            continue

        if (
            validate_required_content
            and _is_true(row.get("required"))
            and not _text(row.get("response"))
         ):
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="overview.required_response_missing",
                    location=f"Workflow Overview.{section}",
                    message=f"Required workflow overview section '{section}' has no response.",
                )
            )


def _validate_workflow_steps(
    rows: list[dict[str, Any]],
    issues: list[PacketValidationIssue],
) -> set[str]:
    step_ids: set[str] = set()

    if not rows:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="steps.no_rows",
                location="Workflow Steps",
                message="Workflow Steps must contain at least one workflow step.",
            )
        )
        return step_ids

    for row_index, row in enumerate(rows, start=2):
        location = f"Workflow Steps row {row_index}"

        step_id = _text(row.get("step_id"))
        sequence = _text(row.get("sequence"))

        if not step_id:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="steps.step_id_missing",
                    location=location,
                    message="step_id is required.",
                )
            )
        elif step_id in step_ids:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="steps.step_id_duplicate",
                    location=location,
                    message=f"Duplicate step_id '{step_id}'.",
                )
            )
        else:
            step_ids.add(step_id)

        _require_text(row, "step_name", location, issues)
        _require_text(row, "owner_role", location, issues)
        _require_text(row, "activity", location, issues)
        _require_text(row, "output", location, issues)

        if not sequence:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="steps.sequence_missing",
                    location=location,
                    message="sequence is required.",
                )
            )
        elif not _is_number(sequence):
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="steps.sequence_invalid",
                    location=location,
                    message=f"sequence must be numeric. Found '{sequence}'.",
                )
            )

    return step_ids


def _validate_policy_controls(
    *,
    rows: list[dict[str, Any]],
    valid_step_ids: set[str],
    issues: list[PacketValidationIssue],
) -> None:
    control_ids: set[str] = set()

    if not rows:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="controls.no_rows",
                location="Policy Controls",
                message="No policy/control rows were provided. Governance recommendations may be incomplete.",
            )
        )
        return

    for row_index, row in enumerate(rows, start=2):
        location = f"Policy Controls row {row_index}"

        control_id = _text(row.get("control_id"))
        control_type = _text(row.get("control_type"))
        approval_required = _text(row.get("approval_required"))
        write_action_allowed = _text(row.get("write_action_allowed"))

        if not control_id:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="controls.control_id_missing",
                    location=location,
                    message="control_id is required.",
                )
            )
        elif control_id in control_ids:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="controls.control_id_duplicate",
                    location=location,
                    message=f"Duplicate control_id '{control_id}'.",
                )
            )
        else:
            control_ids.add(control_id)

        _require_text(row, "control_name", location, issues)
        _require_text(row, "requirement", location, issues)

        if control_type and control_type not in CONTROL_TYPES:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="controls.control_type_invalid",
                    location=location,
                    message=f"control_type '{control_type}' is not an approved value.",
                )
            )

        _validate_step_references(
            value=row.get("applies_to_steps"),
            valid_step_ids=valid_step_ids,
            location=f"{location}.applies_to_steps",
            issues=issues,
        )

        _validate_boolean_value(
            field_name="approval_required",
            value=approval_required,
            location=location,
            issues=issues,
        )
        _validate_boolean_value(
            field_name="write_action_allowed",
            value=write_action_allowed,
            location=location,
            issues=issues,
        )

        if _is_true(approval_required) and not _text(row.get("approval_role")):
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="controls.approval_role_missing",
                    location=location,
                    message="approval_role is required when approval_required=true.",
                )
            )


def _validate_data_dictionary(
    *,
    rows: list[dict[str, Any]],
    valid_step_ids: set[str],
    issues: list[PacketValidationIssue],
) -> set[str]:
    field_names: set[str] = set()

    if not rows:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="dictionary.no_rows",
                location="Data Dictionary",
                message="Data Dictionary must contain at least one field definition.",
            )
        )
        return field_names

    for row_index, row in enumerate(rows, start=2):
        location = f"Data Dictionary row {row_index}"

        field_name = _text(row.get("field_name"))
        data_category = _text(row.get("data_category"))

        if not field_name:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="dictionary.field_name_missing",
                    location=location,
                    message="field_name is required.",
                )
            )
        elif field_name in field_names:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="dictionary.field_name_duplicate",
                    location=location,
                    message=f"Duplicate field_name '{field_name}'.",
                )
            )
        else:
            field_names.add(field_name)

        _require_text(row, "business_meaning", location, issues)
        _require_text(row, "source_system", location, issues)

        if data_category not in DATA_CATEGORIES:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="dictionary.data_category_invalid",
                    location=location,
                    message=f"data_category '{data_category}' is not an approved value.",
                )
            )

        for boolean_field in [
            "required_for_workflow",
            "model_context_allowed",
            "redaction_required",
        ]:
            _validate_boolean_value(
                field_name=boolean_field,
                value=row.get(boolean_field),
                location=location,
                issues=issues,
            )

        if data_category in {"PII", "Credential/Secret", "Regulated"} and _is_true(
            row.get("model_context_allowed")
        ):
            issues.append(
                PacketValidationIssue(
                    severity="warning",
                    code="dictionary.sensitive_field_model_allowed",
                    location=location,
                    message=(
                        f"Field '{field_name}' has sensitive category '{data_category}' "
                        "but model_context_allowed=true. Confirm this is intentional."
                    ),
                )
            )

        _validate_step_references(
            value=row.get("used_in_steps"),
            valid_step_ids=valid_step_ids,
            location=f"{location}.used_in_steps",
            issues=issues,
            allow_blank=True,
        )

    return field_names


def _validate_sample_records(
    *,
    rows: list[dict[str, Any]],
    valid_data_fields: set[str],
    issues: list[PacketValidationIssue],
) -> None:
    if not rows:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="sample_records.no_rows",
                location="Sample Records",
                message="No sample records were provided. Sample-record pattern analysis will be limited.",
            )
        )
        return

    record_ids: set[str] = set()

    sample_headers = set(rows[0].keys()) if rows else set()
    missing_from_dictionary = sample_headers - valid_data_fields

    for header in sorted(missing_from_dictionary):
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="sample_records.column_not_in_dictionary",
                location=f"Sample Records.{header}",
                message=f"Sample Records column '{header}' is missing from Data Dictionary.",
            )
        )

    for row_index, row in enumerate(rows, start=2):
        location = f"Sample Records row {row_index}"

        record_id = _text(row.get("record_id"))

        if not record_id:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="sample_records.record_id_missing",
                    location=location,
                    message="record_id is required.",
                )
            )
        elif record_id in record_ids:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="sample_records.record_id_duplicate",
                    location=location,
                    message=f"Duplicate record_id '{record_id}'.",
                )
            )
        else:
            record_ids.add(record_id)

        for field_name, value in row.items():
            text_value = _text(value)

            if text_value and SECRET_OR_SENSITIVE_PATTERN.search(text_value):
                issues.append(
                    PacketValidationIssue(
                        severity="error",
                        code="sample_records.possible_secret",
                        location=f"{location}.{field_name}",
                        message=(
                            f"Field '{field_name}' appears to contain a secret, token, "
                            "password, or credential-like value."
                        ),
                    )
                )

    if len(rows) < 3:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="sample_records.low_row_count",
                location="Sample Records",
                message="Fewer than 3 sample records were provided. Pattern analysis may be weak.",
            )
        )

    if len(rows) > 50:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="sample_records.high_row_count",
                location="Sample Records",
                message="More than 50 sample records were provided. v1 packets should use representative samples, not data dumps.",
            )
        )


def _validate_target_systems(
    *,
    rows: list[dict[str, Any]],
    issues: list[PacketValidationIssue],
) -> None:
    for row_index, row in enumerate(rows, start=2):
        location = f"Target Systems row {row_index}"

        system_type = _text(row.get("system_type"))

        _require_text(row, "system_name", location, issues)

        if system_type and system_type not in SYSTEM_TYPES:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="systems.system_type_invalid",
                    location=location,
                    message=f"system_type '{system_type}' is not an approved value.",
                )
            )

        _validate_boolean_value(
            field_name="read_access_possible",
            value=row.get("read_access_possible"),
            location=location,
            issues=issues,
        )
        _validate_boolean_value(
            field_name="write_access_possible",
            value=row.get("write_access_possible"),
            location=location,
            issues=issues,
        )


def _read_sheet_rows(
    workbook: Workbook,
    sheet_name: str,
    issues: list[PacketValidationIssue],
) -> list[dict[str, Any]]:
    sheet = workbook[sheet_name]
    headers = _read_headers(sheet)

    if not headers:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="sheet.headers_missing",
                location=sheet_name,
                message=f"Sheet '{sheet_name}' does not have a header row.",
            )
        )
        return []

    rows: list[dict[str, Any]] = []

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(excel_row[: len(headers)])

        if all(_text(value) == "" for value in values):
            continue

        row = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
        }
        rows.append(row)

    return rows


def _read_headers(sheet: Any) -> list[str]:
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [_text(value) for value in raw_headers if _text(value)]


def _validate_boolean_value(
    *,
    field_name: str,
    value: Any,
    location: str,
    issues: list[PacketValidationIssue],
) -> None:
    text = _text(value).lower()

    if not text:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="boolean.required",
                location=f"{location}.{field_name}",
                message=f"{field_name} must be true or false.",
            )
        )
        return

    if text not in {"true", "false"}:
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="boolean.invalid",
                location=f"{location}.{field_name}",
                message=f"{field_name} must be true or false. Found '{value}'.",
            )
        )


def _validate_step_references(
    *,
    value: Any,
    valid_step_ids: set[str],
    location: str,
    issues: list[PacketValidationIssue],
    allow_blank: bool = False,
) -> None:
    refs = _split_semicolon_list(value)

    if not refs and allow_blank:
        return

    if not refs:
        issues.append(
            PacketValidationIssue(
                severity="warning",
                code="steps.reference_blank",
                location=location,
                message="No workflow step reference was provided.",
            )
        )
        return

    for ref in refs:
        if ref not in valid_step_ids:
            issues.append(
                PacketValidationIssue(
                    severity="error",
                    code="steps.reference_invalid",
                    location=location,
                    message=f"Referenced step_id '{ref}' does not exist in Workflow Steps.",
                )
            )


def _require_text(
    row: dict[str, Any],
    field_name: str,
    location: str,
    issues: list[PacketValidationIssue],
) -> None:
    if not _text(row.get(field_name)):
        issues.append(
            PacketValidationIssue(
                severity="error",
                code="field.required",
                location=f"{location}.{field_name}",
                message=f"{field_name} is required.",
            )
        )


def _has_blocking_sheet_errors(issues: list[PacketValidationIssue]) -> bool:
    return any(issue.code == "packet.required_sheet_missing" for issue in issues)


def _is_true(value: Any) -> bool:
    return _text(value).lower() == "true"


def _is_number(value: Any) -> bool:
    try:
        float(_text(value))
        return True
    except ValueError:
        return False


def _split_semicolon_list(value: Any) -> list[str]:
    return [item.strip() for item in _text(value).split(";") if item.strip()]


def _text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _result(
    packet_path: Path,
    issues: list[PacketValidationIssue],
) -> PacketValidationResult:
    error_count = sum(1 for issue in issues if issue.severity == "error")
    warning_count = sum(1 for issue in issues if issue.severity == "warning")

    return PacketValidationResult(
        path=str(packet_path),
        valid=error_count == 0,
        error_count=error_count,
        warning_count=warning_count,
        issues=issues,
    )