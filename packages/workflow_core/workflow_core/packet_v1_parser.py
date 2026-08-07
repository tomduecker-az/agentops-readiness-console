from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from workflow_core.packet_v1_models import (
    DataFieldV1,
    PolicyControlV1,
    TargetSystemV1,
    WorkflowPacketV1,
    WorkflowStepV1,
)
from workflow_core.packet_v1_validator import validate_workflow_packet_v1


PACKET_VERSION = "workflow_packet_v1"

SHEET_WORKFLOW_OVERVIEW = "Workflow Overview"
SHEET_WORKFLOW_STEPS = "Workflow Steps"
SHEET_POLICY_CONTROLS = "Policy Controls"
SHEET_DATA_DICTIONARY = "Data Dictionary"
SHEET_SAMPLE_RECORDS = "Sample Records"
SHEET_GOALS_METRICS = "Goals & Metrics"
SHEET_TARGET_SYSTEMS = "Target Systems"


class WorkflowPacketParseError(ValueError):
    pass


def parse_workflow_packet_v1(
    path: str | Path,
    *,
    validate_before_parse: bool = True,
) -> WorkflowPacketV1:
    packet_path = Path(path)

    if validate_before_parse:
        validation_result = validate_workflow_packet_v1(packet_path)

        if not validation_result.valid:
            issue_summary = "; ".join(
                f"{issue.code}: {issue.location}" for issue in validation_result.issues[:5]
            )
            raise WorkflowPacketParseError(
                "Workflow packet failed validation and cannot be parsed. "
                f"errors={validation_result.error_count}; warnings={validation_result.warning_count}; "
                f"first_issues={issue_summary}"
            )

    workbook = load_workbook(packet_path, read_only=True, data_only=True)

    overview_rows = _read_sheet_rows(workbook, SHEET_WORKFLOW_OVERVIEW)
    overview = _rows_to_key_value_dict(
        rows=overview_rows,
        key_field="section",
        value_field="response",
    )

    workflow_name = overview.get("Workflow Name") or packet_path.stem
    workflow_id = _slugify(workflow_name)

    workflow_steps = _parse_workflow_steps(
        _read_sheet_rows(workbook, SHEET_WORKFLOW_STEPS)
    )
    policy_controls = _parse_policy_controls(
        _read_sheet_rows(workbook, SHEET_POLICY_CONTROLS)
    )
    data_dictionary = _parse_data_dictionary(
        _read_sheet_rows(workbook, SHEET_DATA_DICTIONARY)
    )
    sample_records = _read_sheet_rows(workbook, SHEET_SAMPLE_RECORDS)

    goals_metrics = (
        _rows_to_key_value_dict(
            rows=_read_sheet_rows(workbook, SHEET_GOALS_METRICS),
            key_field="section",
            value_field="response",
        )
        if SHEET_GOALS_METRICS in workbook.sheetnames
        else {}
    )

    target_systems = (
        _parse_target_systems(_read_sheet_rows(workbook, SHEET_TARGET_SYSTEMS))
        if SHEET_TARGET_SYSTEMS in workbook.sheetnames
        else []
    )

    metadata = {
        "source_path": str(packet_path),
        "source_file_name": packet_path.name,
        "sheet_names": list(workbook.sheetnames),
        "workflow_step_count": len(workflow_steps),
        "policy_control_count": len(policy_controls),
        "data_field_count": len(data_dictionary),
        "sample_record_count": len(sample_records),
        "target_system_count": len(target_systems),
        "has_goals_metrics": bool(goals_metrics),
    }

    return WorkflowPacketV1(
        packet_version=PACKET_VERSION,
        workflow_id=workflow_id,
        workflow_name=workflow_name,
        overview=overview,
        workflow_steps=workflow_steps,
        policy_controls=policy_controls,
        data_dictionary=data_dictionary,
        sample_records=sample_records,
        goals_metrics=goals_metrics,
        target_systems=target_systems,
        metadata=metadata,
    )


def _parse_workflow_steps(rows: list[dict[str, Any]]) -> list[WorkflowStepV1]:
    return [
        WorkflowStepV1(
            step_id=_text(row.get("step_id")),
            step_name=_text(row.get("step_name")),
            sequence=_number(row.get("sequence")),
            owner_role=_text(row.get("owner_role")),
            trigger_or_input=_text(row.get("trigger_or_input")),
            activity=_text(row.get("activity")),
            decision_or_rule=_text(row.get("decision_or_rule")),
            systems_used=_split_semicolon_list(row.get("systems_used")),
            data_used=_split_semicolon_list(row.get("data_used")),
            output=_text(row.get("output")),
            exceptions_or_escalations=_text(row.get("exceptions_or_escalations")),
            current_pain_points=_text(row.get("current_pain_points")),
        )
        for row in rows
    ]


def _parse_policy_controls(rows: list[dict[str, Any]]) -> list[PolicyControlV1]:
    return [
        PolicyControlV1(
            control_id=_text(row.get("control_id")),
            control_name=_text(row.get("control_name")),
            control_type=_text(row.get("control_type")),
            applies_to_steps=_split_semicolon_list(row.get("applies_to_steps")),
            requirement=_text(row.get("requirement")),
            approval_required=_bool(row.get("approval_required")),
            approval_role=_text(row.get("approval_role")),
            evidence_required=_text(row.get("evidence_required")),
            write_action_allowed=_bool(row.get("write_action_allowed")),
            retention_requirement=_text(row.get("retention_requirement")),
            source_reference=_text(row.get("source_reference")),
        )
        for row in rows
    ]


def _parse_data_dictionary(rows: list[dict[str, Any]]) -> list[DataFieldV1]:
    return [
        DataFieldV1(
            field_name=_text(row.get("field_name")),
            business_meaning=_text(row.get("business_meaning")),
            source_system=_text(row.get("source_system")),
            data_category=_text(row.get("data_category")),
            required_for_workflow=_bool(row.get("required_for_workflow")),
            model_context_allowed=_bool(row.get("model_context_allowed")),
            redaction_required=_bool(row.get("redaction_required")),
            allowed_values=_split_semicolon_list(row.get("allowed_values")),
            used_in_steps=_split_semicolon_list(row.get("used_in_steps")),
            notes=_text(row.get("notes")),
        )
        for row in rows
    ]


def _parse_target_systems(rows: list[dict[str, Any]]) -> list[TargetSystemV1]:
    return [
        TargetSystemV1(
            system_name=_text(row.get("system_name")),
            system_type=_text(row.get("system_type")),
            read_access_possible=_bool(row.get("read_access_possible")),
            write_access_possible=_bool(row.get("write_access_possible")),
            owner_role=_text(row.get("owner_role")),
            authentication_method=_text(row.get("authentication_method")),
            notes=_text(row.get("notes")),
        )
        for row in rows
    ]


def _read_sheet_rows(workbook: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []

    sheet = workbook[sheet_name]
    headers = _read_headers(sheet)

    rows: list[dict[str, Any]] = []

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(excel_row[: len(headers)])

        if all(_text(value) == "" for value in values):
            continue

        row = {
            header: values[index] if index < len(values) else ""
            for index, header in enumerate(headers)
        }
        rows.append(row)

    return rows


def _read_headers(sheet: Any) -> list[str]:
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [_text(value) for value in raw_headers if _text(value)]


def _rows_to_key_value_dict(
    *,
    rows: list[dict[str, Any]],
    key_field: str,
    value_field: str,
) -> dict[str, str]:
    result: dict[str, str] = {}

    for row in rows:
        key = _text(row.get(key_field))
        value = _text(row.get(value_field))

        if key:
            result[key] = value

    return result


def _split_semicolon_list(value: Any) -> list[str]:
    return [item.strip() for item in _text(value).split(";") if item.strip()]


def _bool(value: Any) -> bool:
    return _text(value).lower() == "true"


def _number(value: Any) -> int | float:
    text = _text(value)

    if not text:
        return 0

    numeric_value = float(text)

    if numeric_value.is_integer():
        return int(numeric_value)

    return numeric_value


def _text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "workflow_packet"