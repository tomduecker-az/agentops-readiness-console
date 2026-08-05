from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DEFAULT = "examples/templates/workflow_packet_v1.xlsx"
EXAMPLE_ACCESS_REQUEST_REVIEW_DEFAULT = (
    "examples/templates/workflow_packet_v1_example_access_request_review.xlsx"
)

SHEETS = {
    "readme": "README",
    "overview": "Workflow Overview",
    "steps": "Workflow Steps",
    "controls": "Policy Controls",
    "dictionary": "Data Dictionary",
    "records": "Sample Records",
    "goals": "Goals & Metrics",
    "systems": "Target Systems",
    "values": "Allowed Values",
}

BOOLEAN_VALUES = ["true", "false"]

DATA_CATEGORIES = [
    "Public",
    "Internal",
    "Confidential",
    "PII",
    "Financial",
    "Security",
    "Sensitive Access",
    "Credential/Secret",
    "Regulated",
    "Derived Metadata",
    "Unknown",
]

CONTROL_TYPES = [
    "approval",
    "audit",
    "data",
    "security",
    "policy",
    "retention",
    "operational",
    "exception",
]

SYSTEM_TYPES = [
    "workflow_system",
    "source_system",
    "execution_system",
    "reporting_system",
    "identity_system",
    "data_store",
    "other",
]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate the Workflow Packet v1 Excel template."
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output .xlsx path. Defaults to the blank template path or example path.",
    )
    parser.add_argument(
        "--example",
        choices=["access_request_review"],
        default=None,
        help="Generate a completed example workbook instead of the blank template.",
    )

    args = parser.parse_args()

    default_output = (
        EXAMPLE_ACCESS_REQUEST_REVIEW_DEFAULT
        if args.example == "access_request_review"
        else OUTPUT_DEFAULT
    )
    output_path = Path(args.output or default_output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    _create_allowed_values_sheet(workbook)
    _create_readme_sheet(workbook)
    _create_workflow_overview_sheet(workbook)
    _create_workflow_steps_sheet(workbook)
    _create_policy_controls_sheet(workbook)
    _create_data_dictionary_sheet(workbook)
    _create_sample_records_sheet(workbook)
    _create_goals_metrics_sheet(workbook)
    _create_target_systems_sheet(workbook)

    workbook[SHEETS["readme"]].sheet_view.showGridLines = False
    workbook[SHEETS["overview"]].sheet_view.showGridLines = False
    workbook[SHEETS["goals"]].sheet_view.showGridLines = False

    if args.example == "access_request_review":
        _populate_access_request_review_example(workbook)

    workbook.save(output_path)

    print("Workflow Packet v1 template generated")
    print(f"- output: {output_path}")


def _create_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["readme"])

    rows = [
        ["Workflow Packet v1 Template"],
        [""],
        ["Purpose"],
        [
            "Use this workbook to provide structured workflow information for an AI Agent Readiness Assessment."
        ],
        [""],
        ["How to use this workbook"],
        ["1. Complete each required sheet."],
        ["2. Use anonymized or synthetic sample records only."],
        ["3. Do not include passwords, tokens, secrets, production identifiers, or regulated personal data."],
        ["4. Keep headers unchanged."],
        ["5. Upload the completed workbook into the assessment app."],
        [""],
        ["Required sheets"],
        ["- Workflow Overview"],
        ["- Workflow Steps"],
        ["- Policy Controls"],
        ["- Data Dictionary"],
        ["- Sample Records"],
        [""],
        ["Optional sheets"],
        ["- Goals & Metrics"],
        ["- Target Systems"],
        [""],
        ["Important"],
        [
            "The application validates this workbook before analysis. Errors block analysis. Warnings lower confidence or limit recommendations."
        ],
    ]

    for row in rows:
        sheet.append(row)

    sheet["A1"].font = Font(size=18, bold=True)
    sheet["A3"].font = Font(bold=True)
    sheet["A6"].font = Font(bold=True)
    sheet["A13"].font = Font(bold=True)
    sheet["A19"].font = Font(bold=True)
    sheet["A23"].font = Font(bold=True)

    sheet.column_dimensions["A"].width = 120
    for row in range(1, sheet.max_row + 1):
        sheet[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")


def _create_workflow_overview_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["overview"])

    headers = ["section", "response", "required", "guidance"]
    rows = [
        [
            "Workflow Name",
            "Access Request Review",
            "true",
            "Name of the workflow being assessed.",
        ],
        [
            "Business Purpose",
            "",
            "true",
            "Why this workflow exists and what business outcome it supports.",
        ],
        [
            "Workflow Trigger",
            "",
            "true",
            "What starts the workflow.",
        ],
        [
            "Workflow Completion Criteria",
            "",
            "true",
            "When the workflow is considered complete.",
        ],
        [
            "Primary Participants",
            "",
            "true",
            "Roles involved in the workflow. Semicolon-separated list is acceptable.",
        ],
        [
            "Systems Involved",
            "",
            "true",
            "Systems used, reviewed, or updated. Semicolon-separated list is acceptable.",
        ],
        [
            "Current Pain Points",
            "",
            "true",
            "Delays, rework, manual effort, audit issues, escalation problems, or quality issues.",
        ],
        [
            "AI Goals",
            "",
            "true",
            "What the organization hopes AI could improve.",
        ],
        [
            "AI No-Go Areas",
            "",
            "true",
            "Actions that should not be automated or should require human approval.",
        ],
        [
            "Known Constraints",
            "",
            "true",
            "Compliance, data, policy, technical, or organizational constraints.",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 70, 14, 80])
    _add_boolean_validation(sheet, "C2:C200")


def _create_workflow_steps_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["steps"])

    headers = [
        "step_id",
        "step_name",
        "sequence",
        "owner_role",
        "trigger_or_input",
        "activity",
        "decision_or_rule",
        "systems_used",
        "data_used",
        "output",
        "exceptions_or_escalations",
        "current_pain_points",
    ]

    rows = [
        [
            "STEP-001",
            "Submit request",
            1,
            "Requesting role",
            "Need for workflow action",
            "Submit request with required information",
            "Request must include required fields",
            "Workflow system",
            "request_type;request_reason",
            "Submitted request",
            "Missing information returned for clarification",
            "Incomplete requests cause rework",
        ],
        [
            "STEP-002",
            "Review request",
            2,
            "Analyst role",
            "Submitted request",
            "Review request for completeness and eligibility",
            "Request must meet documented criteria",
            "Workflow system;source system",
            "request_type;verification_status",
            "Review result",
            "Conflicting or missing information escalated",
            "Manual review consumes analyst time",
        ],
        [
            "STEP-003",
            "Approve or reject",
            3,
            "Approver role",
            "Completed review",
            "Make accountable approval or rejection decision",
            "Approval must be made by authorized reviewer",
            "Workflow system",
            "review_result;approval_status",
            "Approval outcome",
            "Unauthorized or unclear approval routed for resolution",
            "Approval evidence may be inconsistent",
        ],
        [
            "STEP-004",
            "Execute approved action",
            4,
            "Operations role",
            "Approved request",
            "Execute the approved workflow action",
            "Action cannot occur before approval is recorded",
            "Operational system;workflow system",
            "approval_status;action_details",
            "Completed action",
            "Failed execution or missing evidence escalated",
            "Manual execution and updates create audit risk",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 28, 12, 24, 34, 46, 48, 34, 34, 32, 48, 48])


def _create_policy_controls_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["controls"])

    headers = [
        "control_id",
        "control_name",
        "control_type",
        "applies_to_steps",
        "requirement",
        "approval_required",
        "approval_role",
        "evidence_required",
        "write_action_allowed",
        "retention_requirement",
        "source_reference",
    ]

    rows = [
        [
            "CTRL-001",
            "Human approval before consequential action",
            "approval",
            "STEP-003;STEP-004",
            "Consequential actions require recorded human approval before execution",
            "true",
            "Authorized Approver",
            "Approval record",
            "true",
            "Retain according to audit policy",
            "Policy section or source reference",
        ],
        [
            "CTRL-002",
            "No restricted data in model context",
            "data",
            "STEP-001;STEP-002",
            "Restricted fields must be excluded or transformed before model use",
            "false",
            "",
            "Redaction or filtering log",
            "false",
            "Retain model input audit log",
            "AI data handling policy",
        ],
        [
            "CTRL-003",
            "Audit write actions",
            "audit",
            "STEP-004",
            "Workflow writes must be logged with actor action timestamp and approval reference",
            "true",
            "Operations Approver",
            "Audit event",
            "true",
            "Retain according to audit policy",
            "Audit policy",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 42, 18, 24, 62, 18, 28, 32, 22, 34, 34])
    _add_list_validation(sheet, "C2:C500", CONTROL_TYPES)
    _add_boolean_validation(sheet, "F2:F500")
    _add_boolean_validation(sheet, "I2:I500")


def _create_data_dictionary_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["dictionary"])

    headers = [
        "field_name",
        "business_meaning",
        "source_system",
        "data_category",
        "required_for_workflow",
        "model_context_allowed",
        "redaction_required",
        "allowed_values",
        "used_in_steps",
        "notes",
    ]

    rows = [
        [
            "record_id",
            "Unique anonymized sample record identifier",
            "Sample Data",
            "Internal",
            "true",
            "true",
            "false",
            "",
            "STEP-001",
            "Use anonymized values only",
        ],
        [
            "request_type",
            "Type or category of request",
            "Workflow System",
            "Internal",
            "true",
            "true",
            "false",
            "standard;exception;urgent",
            "STEP-001;STEP-002",
            "Allowed if normalized",
        ],
        [
            "request_reason",
            "Business reason for the request",
            "Workflow System",
            "Confidential",
            "true",
            "false",
            "true",
            "",
            "STEP-001;STEP-002",
            "Do not include sensitive free text in model context unless approved",
        ],
        [
            "verification_status",
            "Result of external verification",
            "Source System",
            "Derived Metadata",
            "true",
            "true",
            "false",
            "verified;failed;missing;conflicting",
            "STEP-002",
            "Preferred model-safe derived signal",
        ],
        [
            "approval_status",
            "Current approval state",
            "Workflow System",
            "Internal",
            "true",
            "true",
            "false",
            "pending;approved;rejected;needs_clarification",
            "STEP-003;STEP-004",
            "Must not be treated as sufficient for execution without reviewer validation",
        ],
        [
            "action_details",
            "Details of action to execute",
            "Operational System",
            "Confidential",
            "true",
            "false",
            "true",
            "",
            "STEP-004",
            "May require redaction or derived representation",
        ],
        [
            "current_status",
            "Current workflow status",
            "Workflow System",
            "Internal",
            "true",
            "true",
            "false",
            "submitted;in_review;approved;completed;returned;escalated",
            "STEP-001;STEP-002;STEP-003;STEP-004",
            "Use controlled status values",
        ],
        [
            "notes",
            "Free-text notes or comments",
            "Workflow System",
            "Confidential",
            "false",
            "false",
            "true",
            "",
            "STEP-001;STEP-002;STEP-003;STEP-004",
            "Free text requires careful handling",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [22, 46, 28, 24, 22, 24, 20, 42, 34, 60])
    _add_list_validation(sheet, "D2:D1000", DATA_CATEGORIES)
    _add_boolean_validation(sheet, "E2:G1000")


def _create_sample_records_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["records"])

    headers = [
        "record_id",
        "request_type",
        "request_reason",
        "verification_status",
        "approval_status",
        "action_details",
        "current_status",
        "notes",
    ]

    rows = [
        [
            "REC-001",
            "standard",
            "Standard business request",
            "verified",
            "pending",
            "Standard operational action",
            "in_review",
            "No exception noted",
        ],
        [
            "REC-002",
            "exception",
            "Exception request requiring additional review",
            "verified",
            "pending",
            "Restricted operational action",
            "in_review",
            "Requires additional reviewer",
        ],
        [
            "REC-003",
            "urgent",
            "Urgent request with incomplete information",
            "missing",
            "needs_clarification",
            "Urgent operational action",
            "returned",
            "Missing required supporting details",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 20, 44, 24, 24, 38, 24, 48])


def _create_goals_metrics_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["goals"])

    headers = ["section", "response", "required", "guidance"]

    rows = [
        [
            "Business Goals",
            "",
            "false",
            "What improvement would matter to the organization?",
        ],
        [
            "Current Baselines",
            "",
            "false",
            "Known baseline metrics, such as handling time, clarification rate, rework rate, SLA breach rate, or report preparation time.",
        ],
        [
            "Target Improvements",
            "",
            "false",
            "Desired improvements or pilot targets.",
        ],
        [
            "Risk Tolerance",
            "",
            "false",
            "Errors or risks that are unacceptable.",
        ],
        [
            "Pilot Constraints",
            "",
            "false",
            "Time, team, system, data, compliance, or policy constraints.",
        ],
        [
            "Success Criteria",
            "",
            "false",
            "What would make the pilot worth expanding?",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 70, 14, 90])
    _add_boolean_validation(sheet, "C2:C200")


def _create_target_systems_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["systems"])

    headers = [
        "system_name",
        "system_type",
        "read_access_possible",
        "write_access_possible",
        "owner_role",
        "authentication_method",
        "notes",
    ]

    rows = [
        [
            "Workflow System",
            "workflow_system",
            "true",
            "true",
            "Workflow Owner",
            "OAuth or SSO",
            "Write actions require approval",
        ],
        [
            "Source System",
            "source_system",
            "true",
            "false",
            "System Owner",
            "SSO",
            "Read-only verification source",
        ],
        [
            "Operational System",
            "execution_system",
            "true",
            "true",
            "Operations Owner",
            "Service account or approved integration",
            "Execution requires approval",
        ],
    ]

    _write_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 24, 22, 22, 28, 34, 48])
    _add_list_validation(sheet, "B2:B500", SYSTEM_TYPES)
    _add_boolean_validation(sheet, "C2:D500")


def _create_allowed_values_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(SHEETS["values"])

    columns = {
        "A": ("boolean", BOOLEAN_VALUES),
        "B": ("data_category", DATA_CATEGORIES),
        "C": ("control_type", CONTROL_TYPES),
        "D": ("system_type", SYSTEM_TYPES),
    }

    for col, (header, values) in columns.items():
        sheet[f"{col}1"] = header
        sheet[f"{col}1"].font = Font(bold=True)

        for index, value in enumerate(values, start=2):
            sheet[f"{col}{index}"] = value

        sheet.column_dimensions[col].width = max(18, len(header) + 4)

    sheet.sheet_state = "hidden"


def _write_table(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.append(headers)

    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row_number in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = 36 if row_number == 1 else 54


def _set_column_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = width


def _add_boolean_validation(sheet, cell_range: str) -> None:
    _add_list_validation(sheet, cell_range, BOOLEAN_VALUES)


def _add_list_validation(sheet, cell_range: str, values: list[str]) -> None:
    joined_values = ",".join(values)

    validation = DataValidation(
        type="list",
        formula1=f'"{joined_values}"',
        allow_blank=True,
    )
    validation.error = "Select a valid value from the dropdown list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select a value from the allowed list."
    validation.promptTitle = "Allowed values"

    sheet.add_data_validation(validation)
    validation.add(cell_range)

def _populate_access_request_review_example(workbook: Workbook) -> None:
    _replace_workflow_overview_with_access_request_review(workbook)
    _replace_workflow_steps_with_access_request_review(workbook)
    _replace_policy_controls_with_access_request_review(workbook)
    _replace_data_dictionary_with_access_request_review(workbook)
    _replace_sample_records_with_access_request_review(workbook)
    _replace_goals_metrics_with_access_request_review(workbook)
    _replace_target_systems_with_access_request_review(workbook)


def _replace_workflow_overview_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["overview"]]

    headers = ["section", "response", "required", "guidance"]
    rows = [
        [
            "Workflow Name",
            "Access Request Review",
            "true",
            "Name of the workflow being assessed.",
        ],
        [
            "Business Purpose",
            "Ensure employees receive appropriate application access while preserving security, approval accountability, and audit evidence.",
            "true",
            "Why this workflow exists and what business outcome it supports.",
        ],
        [
            "Workflow Trigger",
            "A manager submits an access request for an employee who needs access to an application, role, or system capability.",
            "true",
            "What starts the workflow.",
        ],
        [
            "Workflow Completion Criteria",
            "The request is approved or rejected, approved access is provisioned when applicable, the ticket is updated, and required approval/evidence records are retained.",
            "true",
            "When the workflow is considered complete.",
        ],
        [
            "Primary Participants",
            "Manager; Identity Analyst; Application Owner; Security Reviewer; IT Provisioning Specialist",
            "true",
            "Roles involved in the workflow.",
        ],
        [
            "Systems Involved",
            "Ticketing System; HRIS; Identity Provider; Application Administration Console; Reporting System",
            "true",
            "Systems used, reviewed, or updated.",
        ],
        [
            "Current Pain Points",
            "Manual verification, incomplete requests, inconsistent evidence packets, unclear routing for privileged access, SLA pressure, and audit preparation effort.",
            "true",
            "Delays, rework, manual effort, audit issues, escalation problems, or quality issues.",
        ],
        [
            "AI Goals",
            "Identify missing information earlier, prepare reviewer packets, summarize model-safe validation results, flag requests needing additional review, improve SLA visibility, and improve audit readiness.",
            "true",
            "What the organization hopes AI could improve.",
        ],
        [
            "AI No-Go Areas",
            "AI must not approve or reject access, provision access, update systems of record, send external communications, or bypass required human review.",
            "true",
            "Actions that should not be automated or should require human approval.",
        ],
        [
            "Known Constraints",
            "PII and sensitive access details must be restricted from model context unless transformed into approved derived signals. Privileged access requires human review. Write actions require approval and audit logging.",
            "true",
            "Compliance, data, policy, technical, or organizational constraints.",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 90, 14, 80])
    _add_boolean_validation(sheet, "C2:C200")


def _replace_workflow_steps_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["steps"]]

    headers = [
        "step_id",
        "step_name",
        "sequence",
        "owner_role",
        "trigger_or_input",
        "activity",
        "decision_or_rule",
        "systems_used",
        "data_used",
        "output",
        "exceptions_or_escalations",
        "current_pain_points",
    ]

    rows = [
        [
            "STEP-001",
            "Submit access request",
            1,
            "Manager",
            "Employee needs application access",
            "Submit access request with employee, application, role, access scope, and business justification.",
            "Request must contain required intake fields before review can begin.",
            "Ticketing System",
            "request_type;request_reason;system_requested;access_scope",
            "Submitted access request ticket",
            "Incomplete requests returned for clarification.",
            "Missing or vague intake data causes rework.",
        ],
        [
            "STEP-002",
            "Verify employee and manager",
            2,
            "Identity Analyst",
            "Submitted access request",
            "Verify employee status, manager relationship, role, and employment status against HR source.",
            "Employee must be active and manager relationship must be valid.",
            "Ticketing System;HRIS",
            "employee_status;manager_verified;verification_status",
            "Employee and manager verification result",
            "Conflicting HR data is escalated.",
            "Manual verification takes analyst time.",
        ],
        [
            "STEP-003",
            "Review access characteristics",
            3,
            "Identity Analyst",
            "Verified request",
            "Review whether access is privileged, sensitive, custom, API-based, SSO-related, or security-impacting.",
            "Requests with privileged or sensitive characteristics require additional review.",
            "Ticketing System;Application Administration Console",
            "request_type;contains_privileged_access;access_scope;system_requested",
            "Review classification",
            "Unclear access scope returned for clarification.",
            "Routing criteria may be inconsistently applied.",
        ],
        [
            "STEP-004",
            "Clarify missing or conflicting intake",
            4,
            "Identity Analyst",
            "Missing, unclear, or conflicting information",
            "Return request to manager or requester for clarification.",
            "Clarification is required before approval routing if required data is missing.",
            "Ticketing System",
            "verification_status;role_mapping_status;notes",
            "Clarification request",
            "Repeated clarification loops may occur.",
            "Clarification loops increase cycle time.",
        ],
        [
            "STEP-005",
            "Application owner decision",
            5,
            "Application Owner",
            "Review-ready request",
            "Approve or reject access request based on application ownership criteria.",
            "Only authorized application owners can approve access.",
            "Ticketing System",
            "application_owner_approval;access_scope;request_reason",
            "Application owner approval outcome",
            "Unauthorized or unclear approvals escalated.",
            "Approval evidence may be incomplete or inconsistent.",
        ],
        [
            "STEP-006",
            "Security review",
            6,
            "Security Reviewer",
            "Privileged, sensitive, custom, API, SSO, or security-impacting request",
            "Review additional risk and approve, reject, or request more information.",
            "Security review is required for privileged or sensitive access.",
            "Ticketing System;Identity Provider",
            "security_review_required;security_review_status;contains_privileged_access",
            "Security review outcome",
            "Incomplete risk context returned for clarification.",
            "Security review routing depends on accurate classification.",
        ],
        [
            "STEP-007",
            "Provision approved access",
            7,
            "IT Provisioning Specialist",
            "Required approvals recorded",
            "Provision approved access in identity provider or application administration system.",
            "Provisioning cannot occur until required approvals are recorded.",
            "Identity Provider;Application Administration Console",
            "application_owner_approval;security_review_status;action_details",
            "Provisioned access",
            "Provisioning failures escalated.",
            "Manual provisioning creates delay and risk of inconsistency.",
        ],
        [
            "STEP-008",
            "Update ticket and evidence",
            8,
            "Identity Analyst",
            "Provisioning complete or request rejected",
            "Update ticket status and attach required approval, provisioning, and review evidence.",
            "System-of-record updates must include evidence references.",
            "Ticketing System",
            "current_status;approval_status;notes",
            "Updated ticket and retained evidence",
            "Missing evidence routed back to responsible reviewer or provisioner.",
            "Evidence collection creates audit effort.",
        ],
        [
            "STEP-009",
            "Escalate near-SLA requests",
            9,
            "Identity Analyst",
            "Request approaches SLA threshold",
            "Escalate requests near SLA breach to appropriate owner.",
            "Near-SLA requests require escalation according to operational policy.",
            "Ticketing System;Reporting System",
            "current_status;sla_risk",
            "Escalation notice",
            "Unclear SLA threshold may delay escalation.",
            "SLA management is manual.",
        ],
        [
            "STEP-010",
            "Prepare weekly report and retain evidence",
            10,
            "Identity Analyst",
            "Weekly reporting cycle",
            "Prepare weekly access request report and retain required evidence.",
            "Reports must be reviewed before distribution if they contain sensitive workflow information.",
            "Reporting System;Ticketing System",
            "current_status;request_type;approval_status;security_review_status",
            "Weekly report and retained evidence",
            "Report discrepancies require follow-up.",
            "Reporting preparation consumes analyst time.",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 34, 12, 26, 38, 58, 58, 42, 46, 42, 52, 52])


def _replace_policy_controls_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["controls"]]

    headers = [
        "control_id",
        "control_name",
        "control_type",
        "applies_to_steps",
        "requirement",
        "approval_required",
        "approval_role",
        "evidence_required",
        "write_action_allowed",
        "retention_requirement",
        "source_reference",
    ]

    rows = [
        [
            "CTRL-001",
            "Minimum intake requirements",
            "policy",
            "STEP-001;STEP-004",
            "Access requests must contain required intake fields before review can proceed.",
            "false",
            "",
            "Completed request form",
            "false",
            "Retain with request ticket",
            "Access intake procedure",
        ],
        [
            "CTRL-002",
            "Employee and manager verification",
            "policy",
            "STEP-002",
            "Employee status and manager relationship must be verified before approval routing.",
            "false",
            "",
            "Verification result",
            "false",
            "Retain with request ticket",
            "Identity operations procedure",
        ],
        [
            "CTRL-003",
            "Application owner approval",
            "approval",
            "STEP-005;STEP-007",
            "Application owner approval is required before access can be provisioned.",
            "true",
            "Application Owner",
            "Recorded approval",
            "true",
            "Retain according to audit policy",
            "Access approval policy",
        ],
        [
            "CTRL-004",
            "Security review for privileged or sensitive access",
            "approval",
            "STEP-003;STEP-006;STEP-007",
            "Privileged, sensitive, custom, API, SSO, or security-impacting access requires security review before provisioning.",
            "true",
            "Security Reviewer",
            "Security review outcome",
            "true",
            "Retain according to audit policy",
            "Privileged access policy",
        ],
        [
            "CTRL-005",
            "No restricted data in model context",
            "data",
            "STEP-001;STEP-002;STEP-003;STEP-006",
            "PII and sensitive access details must be excluded or transformed before model use.",
            "false",
            "",
            "Model-context filtering log",
            "false",
            "Retain model input audit log",
            "AI data handling policy",
        ],
        [
            "CTRL-006",
            "Provisioning requires recorded approvals",
            "approval",
            "STEP-007",
            "Provisioning cannot occur until required approval evidence is present.",
            "true",
            "IT Provisioning Lead",
            "Approval and provisioning evidence",
            "true",
            "Retain according to audit policy",
            "Provisioning control procedure",
        ],
        [
            "CTRL-007",
            "System-of-record updates must be audited",
            "audit",
            "STEP-008",
            "Ticket updates must record actor, timestamp, action, and evidence reference.",
            "true",
            "Identity Operations Lead",
            "Audit event",
            "true",
            "Retain according to audit policy",
            "Audit policy",
        ],
        [
            "CTRL-008",
            "Report review and evidence retention",
            "retention",
            "STEP-010",
            "Weekly reports and supporting evidence must be retained and reviewed before distribution if sensitive information is included.",
            "true",
            "Identity Operations Lead",
            "Reviewed report and evidence package",
            "true",
            "Retain according to audit policy",
            "Reporting and evidence retention policy",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 46, 18, 30, 68, 18, 30, 38, 22, 38, 38])
    _add_list_validation(sheet, "C2:C500", CONTROL_TYPES)
    _add_boolean_validation(sheet, "F2:F500")
    _add_boolean_validation(sheet, "I2:I500")


def _replace_data_dictionary_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["dictionary"]]

    headers = [
        "field_name",
        "business_meaning",
        "source_system",
        "data_category",
        "required_for_workflow",
        "model_context_allowed",
        "redaction_required",
        "allowed_values",
        "used_in_steps",
        "notes",
    ]

    rows = [
        ["record_id", "Unique anonymized sample record identifier", "Sample Data", "Internal", "true", "true", "false", "", "STEP-001", "Use anonymized record IDs only"],
        ["request_type", "Type of access request", "Ticketing System", "Internal", "true", "true", "false", "standard;privileged;emergency", "STEP-001;STEP-003;STEP-010", "Allowed if normalized"],
        ["request_reason", "Business justification for requested access", "Ticketing System", "Confidential", "true", "false", "true", "", "STEP-001;STEP-005", "Free text should be summarized or excluded unless approved"],
        ["employee_status", "Employment status from HR source", "HRIS", "PII", "true", "false", "true", "active;inactive;terminated;unknown", "STEP-002", "Use derived verification status in model context"],
        ["manager_verified", "Whether manager relationship was verified", "HRIS", "Derived Metadata", "true", "true", "false", "true;false", "STEP-002", "Model-safe derived signal"],
        ["role_mapping_status", "Whether requested access maps cleanly to an approved role", "Ticketing System", "Derived Metadata", "true", "true", "false", "matched;missing;conflicting", "STEP-003;STEP-004", "Useful for clarification routing"],
        ["contains_privileged_access", "Whether requested access includes privileged access", "Ticketing System", "Sensitive Access", "true", "false", "true", "true;false", "STEP-003;STEP-006", "Use only approved derived handling"],
        ["access_scope", "Scope of requested access", "Ticketing System", "Sensitive Access", "true", "false", "true", "standard;privileged;admin;custom", "STEP-001;STEP-003;STEP-005", "Do not send raw sensitive scope unless approved"],
        ["system_requested", "Application or system for which access is requested", "Ticketing System", "Internal", "true", "true", "false", "", "STEP-001;STEP-003", "Allowed if not itself sensitive"],
        ["verification_status", "Overall verification result", "HRIS", "Derived Metadata", "true", "true", "false", "verified;failed;missing;conflicting", "STEP-002;STEP-004", "Preferred model-safe verification signal"],
        ["application_owner_approval", "Application owner approval state", "Ticketing System", "Internal", "true", "true", "false", "pending;approved;rejected;needs_clarification", "STEP-005;STEP-007;STEP-010", "Approval alone does not authorize provisioning without all required reviews"],
        ["security_review_required", "Whether security review is required", "Ticketing System", "Derived Metadata", "true", "true", "false", "true;false", "STEP-003;STEP-006", "Derived routing signal"],
        ["security_review_status", "Security review state", "Ticketing System", "Internal", "true", "true", "false", "not_required;pending;approved;rejected;needs_clarification", "STEP-006;STEP-007;STEP-010", "Required before provisioning for privileged or sensitive requests"],
        ["approval_status", "Overall approval state", "Ticketing System", "Internal", "true", "true", "false", "pending;approved;rejected;needs_clarification", "STEP-005;STEP-007;STEP-008;STEP-010", "Must be validated against required approvals before execution"],
        ["action_details", "Provisioning action to execute", "Identity Provider", "Confidential", "true", "false", "true", "", "STEP-007", "Do not expose raw provisioning details to model unless approved"],
        ["current_status", "Current workflow status", "Ticketing System", "Internal", "true", "true", "false", "submitted;in_review;pending_security_review;ready_to_provision;completed;returned;escalated", "STEP-008;STEP-009;STEP-010", "Controlled workflow state"],
        ["sla_risk", "Whether request is approaching SLA threshold", "Reporting System", "Derived Metadata", "false", "true", "false", "low;medium;high", "STEP-009", "Useful for prioritization"],
        ["notes", "Free-text workflow notes", "Ticketing System", "Confidential", "false", "false", "true", "", "STEP-004;STEP-008", "Free text requires careful handling"],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 56, 30, 24, 22, 24, 20, 48, 42, 70])
    _add_list_validation(sheet, "D2:D1000", DATA_CATEGORIES)
    _add_boolean_validation(sheet, "E2:G1000")


def _replace_sample_records_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["records"]]

    headers = [
        "record_id",
        "request_type",
        "request_reason",
        "employee_status",
        "manager_verified",
        "role_mapping_status",
        "contains_privileged_access",
        "access_scope",
        "system_requested",
        "verification_status",
        "application_owner_approval",
        "security_review_required",
        "security_review_status",
        "approval_status",
        "action_details",
        "current_status",
        "sla_risk",
        "notes",
    ]

    rows = [
        [
            "AR-1001",
            "standard",
            "Routine business access request",
            "active",
            "true",
            "matched",
            "false",
            "standard",
            "Expense Management",
            "verified",
            "pending",
            "false",
            "not_required",
            "pending",
            "Standard role assignment",
            "in_review",
            "low",
            "Standard request ready for application owner review",
        ],
        [
            "AR-1002",
            "privileged",
            "Privileged access needed for operational support",
            "active",
            "true",
            "matched",
            "true",
            "admin",
            "Customer Operations Console",
            "verified",
            "approved",
            "true",
            "pending",
            "pending",
            "Privileged role assignment",
            "pending_security_review",
            "medium",
            "Requires security review before provisioning",
        ],
        [
            "AR-1003",
            "privileged",
            "Privileged access request with incomplete role mapping",
            "active",
            "true",
            "missing",
            "true",
            "custom",
            "Reporting Administration",
            "missing",
            "needs_clarification",
            "true",
            "needs_clarification",
            "needs_clarification",
            "Custom role assignment",
            "returned",
            "high",
            "Role mapping incomplete and clarification required",
        ],
        [
            "AR-1004",
            "standard",
            "Standard access request for new team member",
            "active",
            "true",
            "matched",
            "false",
            "standard",
            "Document Repository",
            "verified",
            "approved",
            "false",
            "not_required",
            "approved",
            "Standard group assignment",
            "ready_to_provision",
            "medium",
            "Ready to provision after approval evidence check",
        ],
        [
            "AR-1005",
            "emergency",
            "Urgent time-sensitive access request",
            "active",
            "true",
            "conflicting",
            "true",
            "privileged",
            "Incident Response Tool",
            "conflicting",
            "pending",
            "true",
            "pending",
            "pending",
            "Privileged emergency access",
            "escalated",
            "high",
            "Conflicting role mapping and SLA escalation",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [16, 18, 46, 20, 20, 24, 28, 22, 34, 24, 28, 26, 28, 24, 34, 28, 16, 58])


def _replace_goals_metrics_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["goals"]]

    headers = ["section", "response", "required", "guidance"]
    rows = [
        [
            "Business Goals",
            "Reduce manual analyst preparation effort, reduce avoidable clarification loops, improve reviewer packet consistency, improve SLA visibility, and improve audit readiness.",
            "false",
            "What improvement would matter to the organization?",
        ],
        [
            "Current Baselines",
            "Average handling time, clarification rate, SLA-risk count, security-review routing accuracy, reviewer correction rate, and weekly reporting effort are not fully baselined yet.",
            "false",
            "Known baseline metrics.",
        ],
        [
            "Target Improvements",
            "Pilot should show measurable reduction in analyst preparation time and clarification rework without increasing missed deficiencies or unsupported recommendations.",
            "false",
            "Desired improvements or pilot targets.",
        ],
        [
            "Risk Tolerance",
            "No autonomous approval, provisioning, system-of-record updates, report distribution, or restricted data in model context.",
            "false",
            "Errors or risks that are unacceptable.",
        ],
        [
            "Pilot Constraints",
            "Pilot should be read-only/advisory, use model-safe derived fields, preserve human approval authority, and maintain audit evidence.",
            "false",
            "Time, team, system, data, compliance, or policy constraints.",
        ],
        [
            "Success Criteria",
            "Reviewers accept AI-prepared packets as useful, corrections are tracked, no restricted data enters model context, and no required approval/provisioning control is bypassed.",
            "false",
            "What would make the pilot worth expanding?",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [28, 90, 14, 90])
    _add_boolean_validation(sheet, "C2:C200")


def _replace_target_systems_with_access_request_review(workbook: Workbook) -> None:
    sheet = workbook[SHEETS["systems"]]

    headers = [
        "system_name",
        "system_type",
        "read_access_possible",
        "write_access_possible",
        "owner_role",
        "authentication_method",
        "notes",
    ]

    rows = [
        [
            "Ticketing System",
            "workflow_system",
            "true",
            "true",
            "Identity Operations",
            "SSO or OAuth",
            "Ticket updates require approval/audit controls.",
        ],
        [
            "HRIS",
            "source_system",
            "true",
            "false",
            "HR Operations",
            "SSO",
            "Read-only verification source.",
        ],
        [
            "Identity Provider",
            "identity_system",
            "true",
            "true",
            "IT Security",
            "Approved service account",
            "Provisioning requires recorded approval evidence.",
        ],
        [
            "Application Administration Console",
            "execution_system",
            "true",
            "true",
            "Application Owner",
            "SSO or approved integration",
            "Application-specific access execution system.",
        ],
        [
            "Reporting System",
            "reporting_system",
            "true",
            "true",
            "Identity Operations",
            "SSO",
            "Report distribution requires review when sensitive information is included.",
        ],
    ]

    _replace_table(sheet, headers, rows)
    _set_column_widths(sheet, [34, 24, 22, 22, 30, 34, 62])
    _add_list_validation(sheet, "B2:B500", SYSTEM_TYPES)
    _add_boolean_validation(sheet, "C2:D500")


def _replace_table(sheet, headers: list[str], rows: list[list[object]]) -> None:
    sheet.delete_rows(1, sheet.max_row)
    _write_table(sheet, headers, rows)


if __name__ == "__main__":
    main()